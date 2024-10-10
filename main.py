import sys
import time
import logging
import json
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from datetime import datetime, timedelta
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException, TimeoutException, WebDriverException
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import random
import msvcrt  # Windowsの場合
import threading
from colorama import init, Fore, Back, Style
from pyfiglet import Figlet
import time
import sys
import re

init(autoreset=True)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def print_progress(current, total):
    fig = Figlet(font='big')
    progress_text = f"{current}/{total}"
    ascii_art = fig.renderText(progress_text)
    
    sys.stdout.write('\033[H')   # Move cursor to home position
    print('\n' * 10)  # 10行の改行を挿入
    
    print(f"{Fore.YELLOW}{Style.BRIGHT}取得済みポスト数:")
    print(f"{Fore.CYAN}{ascii_art}")
    
    for _ in range(5):  # Blink effect
        sys.stdout.write('\033[?25l')  # Hide cursor
        sys.stdout.write(f"\r{Fore.RED}{Back.WHITE}{Style.BRIGHT}進捗状況: {progress_text}{Style.RESET_ALL}")
        sys.stdout.flush()
        time.sleep(0.3)
        sys.stdout.write(f"\r{' ' * (30 + len(progress_text))}")
        sys.stdout.flush()
        time.sleep(0.3)
    
    sys.stdout.write('\033[?25h')  # Show cursor
    print("\n" + "=" * 50)

def setup_driver():
    chrome_options = Options()
    #chrome_options.add_argument("--headless")
    user_agents = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0 Safari/605.1.15",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0"
    ]
    
    selected_user_agent = random.choice(user_agents)
    chrome_options.add_argument(f"user-agent={selected_user_agent}")
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    logging.info("Chromeドライバーを設定しました。ヘッドレスモード: 有効")
    return driver

def wait_and_log(seconds, message):
    logging.info(f"{message}、{seconds}秒間待機します...")
    time.sleep(seconds)

def key_listener(shared_data):
    while not shared_data['stop']:
        if msvcrt.kbhit():  # Windowsの場合
        # if select.select([sys.stdin], [], [], 0.1)[0]:  # Unixの場合
            key = msvcrt.getch().decode('utf-8').lower()
            if key == 'q':
                shared_data['terminate'] = True
                break

def save_cookies(driver, filename='cookies.json'):
    """ドライバーのクッキーを保存する"""
    with open(filename, 'w') as f:
        json.dump(driver.get_cookies(), f)
    logging.info(f"クッキーを {filename} に保存しました")

def load_cookies(driver, filename='cookies.json'):
    """保存されたクッキーをドライバーにロードする"""
    with open(filename, 'r') as f:
        cookies = json.load(f)
    for cookie in cookies:
        driver.add_cookie(cookie)
    logging.info(f"{filename} からクッキーをロードしました")

def login_to_threads(driver, username, password):
    url = "https://www.threads.net/login/"

    try:
        driver.get(url)
        logging.info(f"ログインページにアクセスしています: {url}")
        
        username_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='text'][class*='x1i10hfl'][class*='x1a2a7pz']"))
        )
        username_field.clear()
        username_field.send_keys(username)
        logging.info("ユーザー名を入力しました")
        
        password_field = driver.find_element(By.CSS_SELECTOR, "input[type='password']")
        password_field.clear()
        password_field.send_keys(password)
        logging.info("パスワードを入力しました")
        
        wait_and_log(5, "ログインボタンをクリックする前に待機")
        
        login_button_xpath = "//div[@role='button' and contains(@class, 'x1i10hfl') and contains(@class, 'x1qjc9v5')]//div[contains(text(), 'Log in') or contains(text(), 'ログイン')]"
        login_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, login_button_xpath))
        )
        driver.execute_script("arguments[0].click();", login_button)
        logging.info("ログインボタンをクリックしました")
        
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        logging.info("ログインが完了し、ページが読み込まれました")

        time.sleep(10)
        
        # ログイン成功後、クッキーを保存
        save_cookies(driver)
        
        return True

    except Exception as e:
        logging.error(f"ログイン中にエラーが発生しました: {e}")


def check_login_status(driver):
    """'Post'または'投稿'要素の存在に基づいてログイン状態を確認する"""
    try:
        # 'Post'または'投稿'要素を探す
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, 
                "//div[contains(@class, 'xc26acl') and contains(@class, 'x6s0dn4') and contains(@class, 'x78zum5') and (contains(text(), 'Post') or contains(text(), '投稿'))]"
            ))
        )
        logging.info("'Post'または'投稿'要素が見つかりました。ログイン状態が確認されました。")
        return True
    except:
        logging.info("'Post'または'投稿'要素が見つかりません。ログアウト状態です。")
        return False

def refresh_session(driver, username, password):
    """セッションを更新する"""
    if not check_login_status(driver):
        logging.info("セッションが切れています。再ログインを試みます。")
        return login_to_threads(driver, username, password)
    return True

def scroll_page(driver):
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    logging.info("ページを下部までスクロールしました")
    time.sleep(2)

def get_post_hrefs(html_content):
    soup = BeautifulSoup(html_content, 'html.parser')
    post_hrefs = []
    elements = soup.find_all('a', class_=['x1i10hfl', 'x1lliihq'], href=True)
    for element in elements:
        href = element['href']
        if '/post/' in href and href not in post_hrefs:
            post_hrefs.append(href)
    return post_hrefs

def access_threads_profile(driver, username, login_username, login_password):
    url = f"https://www.threads.net/@{username}"
    all_post_hrefs = []
    
    try:
        # セッションを更新
        if not refresh_session(driver, login_username, login_password):
            raise Exception("セッションの更新に失敗しました")

        driver.get(url)
        logging.info(f"プロフィールページにアクセスしています: {url}")
        
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        logging.info(f"@{username} のプロフィールページの読み込みが完了しました")
        
        no_new_links_count = 0
        
        while True:
            previous_count = len(all_post_hrefs)
            new_hrefs = get_post_hrefs(driver.page_source)
            all_post_hrefs.extend([href for href in new_hrefs if href not in all_post_hrefs])
            
            if len(all_post_hrefs) > previous_count:
                logging.info(f"新たに {len(all_post_hrefs) - previous_count} 件の投稿リンクを見つけました。合計: {len(all_post_hrefs)} 件")
                no_new_links_count = 0
            else:
                no_new_links_count += 1
                logging.info(f"新しいリンクが見つかりません。スクロール回数: {no_new_links_count}")
            
            if no_new_links_count >= 3:
                logging.info("3回連続で新しいリンクが見つからなかったため、収集を終了します")
                break
            
            scroll_page(driver)
            time.sleep(2)
        
        logging.info("データ収集が完了しました")
        logging.info(f"収集された一意の投稿リンクの総数: {len(all_post_hrefs)} 件")
        
    except Exception as e:
        logging.error(f"プロフィールページのアクセス中にエラーが発生しました: {e}")
        return []
    
    return all_post_hrefs

def extract_post_datetime(driver):
    try:
        # 親要素を特定
        parent_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "span.x1lliihq.x1plvlek.xryxfnj.x1n2onr6.x193iq5w.xeuugli.x1fj9vlw.x13faqbe.x1vvkbs.x1s928wv.xhkezso.x1gmr53x.x1cpjm7i.x1fgarty.x1943h6x.x1i0vuye.xjohtrz.xo1l8bm.x12rw4y6.x1yc453h"))
        )
        
        # time要素を特定
        time_element = parent_element.find_element(By.TAG_NAME, "time")
        
        # datetime属性から投稿日時を取得
        datetime_str = time_element.get_attribute('datetime')
        
        # ISO形式の文字列をdatetimeオブジェクトに変換
        post_datetime = datetime.fromisoformat(datetime_str.replace('Z', '+00:00'))
        
        # 日本時間に変換（UTC+9）
        japan_time = post_datetime + timedelta(hours=9)
        
        # 指定のフォーマットに変換
        formatted_datetime = japan_time.strftime("%m月%d日%H時%M分")
        
        return formatted_datetime
    except Exception as e:
        logging.error(f"投稿日時の抽出中にエラーが発生しました: {e}")
        return None


def extract_like_count(driver):
    """いいね数を取得する"""
    selectors = [
        "div.x6s0dn4.x17zd0t2.x78zum5.xl56j7k span.x17qophe.x10l6tqk.x13vifvy",
        "span.x1lliihq.x1plvlek.xryxfnj.x1n2onr6.x1ji0vk5.x18bv5gf.x193iq5w.xeuugli.x1fj9vlw.x13faqbe.x1vvkbs.x1s928wv.xhkezso.x1gmr53x.x1cpjm7i.x1fgarty.x1943h6x.x1i0vuye.xmd891q.xo1l8bm.xc82ewx.x1yc453h span.x17qophe.x10l6tqk.x13vifvy",
        "div.xu9jpxn.x1n2onr6.xqcsobp.x12w9bfk.x1wsgiic.xuxw1ft.x1bl4301 span.x17qophe.x10l6tqk.x13vifvy"
    ]
    
    for selector in selectors:
        try:
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, selector))
            )
            like_count = element.text.strip()
            
            like_count_int = int(like_count) if like_count.isdigit() else 0
            
            logging.info(f"いいね数を取得しました: {like_count_int}")
            return like_count_int
        except TimeoutException:
            continue
        except NoSuchElementException:
            continue
        except Exception as e:
            logging.error(f"セレクター {selector} での処理中にエラーが発生しました: {e}")
            continue

def extract_comment_count(driver):
    try:
        # コメントアイコンと数字を含む要素を複数取得
        comment_elements = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.x6s0dn4.x17zd0t2.x78zum5.xl56j7k"))
        )

        # 2番目の要素（インデックス1）を親要素として選択
        if len(comment_elements) >= 2:
            parent_element = comment_elements[1]
        else:
            logging.warning("2番目の要素が見つかりません。最初の要素を使用します。")
            parent_element = comment_elements[0]

        # コメント数を取得
        count_element = parent_element.find_element(By.CSS_SELECTOR, "span.x17qophe.x10l6tqk.x13vifvy")
        comment_count = count_element.text.strip()
        
        logging.info(f"抽出されたコメント数: {comment_count}")
        
        return int(comment_count) if comment_count.isdigit() else 0

    except TimeoutException:
        logging.error("コメント要素を見つけるのにタイムアウトしました")
        return 0
    except NoSuchElementException:
        logging.error("コメント要素が見つかりませんでした")
        return 0
    except IndexError:
        logging.error("必要な数の要素が見つかりませんでした")
        return 0
    except Exception as e:
        logging.error(f"コメント数の抽出中に予期せぬエラーが発生しました: {e}")
        return 0
        
def extract_image_urls(driver):
    try:
        # 画像を含む親要素を待機して取得
        parent_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.x1xmf6yo"))
        )

        # 画像要素を探す
        image_elements = parent_element.find_elements(By.CSS_SELECTOR, "img.xl1xv1r")
        
        # 画像URLを取得
        image_urls = [img.get_attribute('src') for img in image_elements if img.get_attribute('src')]
        
        # 画像がない場合は「なし」を返す
        return image_urls if image_urls else ["なし"]
    except Exception as e:
        logging.error(f"画像URLの抽出中にエラーが発生しました: {e}")
        return ["なし"]

def extract_caption(driver):
    try:
        # キャプションを含む親要素を待機して取得
        parent_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.x1a6qonq.x6ikm8r.x10wlt62.xj0a0fe.x126k92a.x6prxxf.x7r5mf7"))
        )
        
        # h1とh2要素を取得
        h_elements = parent_element.find_elements(By.CSS_SELECTOR, "h1, h2")
        
        # テキスト部分を抽出して結合
        caption_text = ' '.join([elem.text.strip() for elem in h_elements])
        
        return caption_text if caption_text else None
    
    except Exception as e:
        logging.error(f"キャプションの抽出中にエラーが発生しました: {e}")
        return None

def extract_impression_count(driver):
    """
    表示数を取得し、多様な形式（通常の数字、'k'、'M'表記、カンマ区切り、日本語の万単位）に対応する
    
    :param driver: Selenium WebDriverインスタンス
    :return: int 表示数（整数値）
    """
    try:
        # ビュー数を含む要素を待機して取得
        view_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, 
                "//div[contains(@class, 'x1b12d3d') and contains(@class, 'x6ikm8r')]//span[contains(@class, 'x1lliihq') and contains(@class, 'x1plvlek')]//span[contains(@class, 'x1lliihq') and contains(@class, 'x193iq5w')]"))
        )
        
        # テキストを取得
        view_text = view_element.text.strip()
        
        # 数字部分のみを抽出するための正規表現
        number_pattern = r'([\d,\.]+)\s*(k|m|万)?'
        match = re.search(number_pattern, view_text, re.IGNORECASE)
        
        if match:
            number, unit = match.groups()
            # カンマを除去し、浮動小数点数に変換
            number = float(number.replace(',', ''))
            
            if unit:
                unit = unit.lower()
                if unit == 'k':
                    number *= 1000
                elif unit == 'm':
                    number *= 1000000
                elif unit == '万':
                    number *= 10000
            
            impression_count = int(number)
            
            logging.info(f"表示数を取得しました: {impression_count} (元の表記: {view_text})")
            return impression_count
        else:
            logging.warning(f"表示数の形式が認識できません: {view_text}")
            return 0
    except Exception as e:
        logging.error(f"表示数の抽出中にエラーが発生しました: {e}")
        raise

def extract_reply_count(driver, username, max_scroll_attempts=5, scroll_pause_time=2):
    """
    指定されたユーザーの投稿に対するコメントの返信数を抽出する関数。
    重複するコメントを排除します。
    
    :param driver: Seleniumのウェブドライバーインスタンス
    :param username: 対象ユーザーのユーザーネーム
    :param max_scroll_attempts: 最大スクロール試行回数（デフォルト: 5）
    :param scroll_pause_time: スクロール後の待機時間（秒）（デフォルト: 2）
    :return: 抽出されたコメントの返信数の合計
    """
    logging.info(f"{username}のコメント返信数の抽出を開始します。")

    def get_comment_count(outer_element):
        """個々のコメント要素から返信数を抽出し、キャプションテキストを返す関数。"""
        try:
            caption_element = outer_element.find_element(By.CSS_SELECTOR, "span.x1lliihq.x1plvlek.xryxfnj.x1n2onr6.x193iq5w.xeuugli.x1fj9vlw.x13faqbe.x1vvkbs.x1s928wv.xhkezso.x1gmr53x.x1cpjm7i.x1fgarty.x1943h6x.x1i0vuye.xjohtrz.xo1l8bm.xp07o12.x1yc453h.xat24cr.xdj266r")
            caption_text = caption_element.text

            reply_spans = outer_element.find_elements(By.CSS_SELECTOR, "span.x1lliihq.x1plvlek.xryxfnj.x1n2onr6.x193iq5w.xeuugli.x1fj9vlw.x13faqbe.x1vvkbs.x1s928wv.xhkezso.x1gmr53x.x1cpjm7i.x1fgarty.x1943h6x.x1i0vuye.xmd891q.xo1l8bm.xc82ewx.x1yc453h")
            
            if len(reply_spans) >= 2:
                reply_count = reply_spans[1].find_element(By.CSS_SELECTOR, "span.x17qophe.x10l6tqk.x13vifvy").text
                return int(reply_count == "1"), caption_text
            return 0, caption_text
        except NoSuchElementException:
            logging.debug("キャプションまたは返信数の要素が見つかりませんでした。")
            return 0, ""
        except Exception as e:
            logging.error(f"コメント情報の取得中に予期せぬエラーが発生しました: {str(e)}")
            return 0, ""

    try:
        comment_count = 0
        processed_comments = set()  # 処理済みコメントを追跡するためのセット

        for scroll_attempt in range(max_scroll_attempts):
            logging.info(f"スクロール試行 {scroll_attempt + 1}/{max_scroll_attempts}")
            
            outer_elements = WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.x78zum5.xdt5ytf"))
            )

            for outer_element in outer_elements[1:]:  # 最初の要素（投稿自体）をスキップ
                count, caption = get_comment_count(outer_element)
                
                if caption not in processed_comments:
                    processed_comments.add(caption)
                    comment_count += count
                    logging.info(f"新しいコメント処理: {caption[:30]}... 返信数: {count}")
                else:
                    logging.debug(f"重複コメントをスキップ: {caption[:30]}...")

            # ページをスクロール
            last_height = driver.execute_script("return document.body.scrollHeight")
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(scroll_pause_time)

            new_height = driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                logging.info("これ以上スクロールできません。処理を終了します。")
                break

        logging.info(f"合計 {comment_count} 件の返信数1のコメントを検出しました。")
        return comment_count

    except TimeoutException:
        logging.warning("ページの読み込みがタイムアウトしました。コメント数は0として扱います。")
        return 0
    except Exception as e:
        logging.error(f"予期せぬエラーが発生しました: {str(e)}")
        return 0

def process_posts(driver, post_hrefs, target_username):
    processed_posts = []
    previous_post = None

    index = 1
    total_posts = len(post_hrefs)

    for href in post_hrefs:
        if shared_data['terminate']:
            logging.info("ユーザーによる中断を検出しました。処理を終了します。")
            break

        try:
            full_url = f"https://www.threads.net{href}"
            driver.get(full_url)
            logging.info(f"アクセス中: {full_url}")
            
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.TAG_NAME, "body"))
                    )
                    break  # 成功したらループを抜ける
                except TimeoutException:
                    if attempt < max_retries - 1:  # 最後の試行でなければ
                        logging.warning(f"ページ読み込みに失敗しました。再試行します。(試行回数: {attempt + 1})")
                        time.sleep(2)  # 短い待機時間を設ける
                    else:
                        logging.info("ページの読み込みが繰り返し失敗しました。")
                        logging.info("ページが読み込まれなかったのでこのページの処理をスキップします。")
                        continue

            logging.info(f"ページが正常に読み込まれました。")

            time.sleep(2)  # 短い待機時間を設ける

            # 429エラーの検出
            if is_rate_limited(driver):
                logging.error("429エラー（Too Many Requests）が検出されました。処理を中断します。")
                break  # ループを抜ける

            # インプレッション数を抽出
            impression_count = extract_impression_count(driver)
            logging.info(f"インプレッション数処理完了")

            # 投稿日時を抽出
            post_datetime = extract_post_datetime(driver)
            logging.info(f"投稿日時処理完了")

            # いいね数を抽出
            like_count = extract_like_count(driver)
            logging.info(f"いいね処理完了")

            # 画像URLを抽出
            image_urls = extract_image_urls(driver)
            logging.info(f"画像URL処理完了")

            # キャプションを抽出
            caption = extract_caption(driver)
            logging.info(f"キャプション処理完了")

            # コメント数を抽出
            comment_count = extract_comment_count(driver)
            logging.info(f"コメント数処理完了")

            reply_count = 0

            if comment_count > 0:
                # コメント返信数を抽出
                reply_count = extract_reply_count(driver, target_username)
                logging.info(f"コメント返信数処理完了")
            
            current_post = {
                'url': full_url,
                'datetime': post_datetime,
                'like_count': like_count,
                'comment_count': comment_count,
                'reply_count': reply_count,
                'impression_count': impression_count, 
                'caption': caption,
                'image_urls': image_urls,
            }
            
            # 前の投稿と比較
            if previous_post and is_duplicate_post(previous_post, current_post):
                logging.info(f"重複投稿をスキップ: {full_url}")
                continue
            
            processed_posts.append(current_post)
            previous_post = current_post
            
            logging.info(f"処理完了: {full_url}")
            logging.info(f"  - 日時: {post_datetime}")
            logging.info(f"  - インプレッション数: {impression_count}")
            logging.info(f"  - いいね数: {like_count}")
            logging.info(f"  - コメント数: {comment_count}")
            logging.info(f"  - コメント返信数: {reply_count}") 
            logging.info(f"  - キャプション: {caption}")
            logging.info(f"  - 画像URL: {', '.join(image_urls)}")

            logging.info(f"--------------------")
            logging.info(f"★★現在取得投稿数：{index}★★")
            logging.info(f"--------------------")
            
            index = index + 1
            
            # レート制限を回避するための待機
            time.sleep(2)
        
        except Exception as e:
            logging.error(f"投稿の処理中にエラーが発生しました: {full_url}, エラー: {e}")

    return processed_posts

def is_rate_limited(driver):
    try:
        # ステータスコードを取得
        status_code = driver.execute_script("return window.performance.timing.responseStart > 0 ? window.performance.getEntries()[0].responseStatus : null;")
        
        if status_code == 429:
            return True

        # エラーメッセージを確認
        error_messages = [
            "Too Many Requests",
            "レート制限を超えました",
            "Rate limit exceeded"
        ]
        
        for message in error_messages:
            if message in driver.page_source:
                return True

        return False
    except WebDriverException:
        # エラーが発生した場合、安全のためにTrueを返す
        return True

def is_duplicate_post(previous_post, current_post):
    """
    前の投稿と現在の投稿が重複しているかどうかを判断する関数
    """
    return (
        previous_post['datetime'] == current_post['datetime'] and
        previous_post['caption'] == current_post['caption']
    )

def validate_input(input_string, input_type):
    if not input_string:
        raise ValueError(f"{input_type}が空です。有効な値を入力してください。")
    return input_string

def save_to_excel(processed_posts, target_username):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{target_username}の投稿データ"

    # ヘッダーの設定
    headers = ['投稿日時', 'キャプション', '画像', 'いいね数', 'コメント数', 'コメント返信数', 'インプレッション数', '投稿URL']
    orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = orange_fill

    # データの書き込み
    row = 2
    for post in processed_posts:
        start_row = row
        ws.cell(row=row, column=1, value=post['datetime'])
        ws.cell(row=row, column=2, value=post['caption'])
        
        # 画像URLを縦に並べてハイパーリンクとして追加
        image_urls = post['image_urls']
        if image_urls != ['なし']:
            for i, url in enumerate(image_urls, start=1):
                cell = ws.cell(row=row+i-1, column=3, value=f'画像{i}')
                cell.hyperlink = url
                cell.font = Font(color="0563C1", underline="single")  # Hyperlink style
            row += len(image_urls) - 1  # 画像の数だけ行を増やす
        else:
            ws.cell(row=row, column=3, value='なし')
        
        ws.cell(row=start_row, column=4, value=post['like_count'])
        ws.cell(row=start_row, column=5, value=post['comment_count'])
        ws.cell(row=start_row, column=6, value=post['reply_count'])
        ws.cell(row=start_row, column=7, value=post['impression_count']) 
        
        # 投稿URLをハイパーリンクとして追加
        cell = ws.cell(row=start_row, column=8, value='投稿リンク')
        cell.hyperlink = post['url']
        cell.font = Font(color="0563C1", underline="single")  # Hyperlink style

        # セルの結合
        for col in [1, 2, 4, 5, 6, 7, 8]:
            if row > start_row:
                ws.merge_cells(start_row=start_row, start_column=col, end_row=row, end_column=col)
                ws.cell(row=start_row, column=col).alignment = Alignment(vertical='center')

        row += 1  # 次の投稿のために行を進める

    # 列幅の自動調整
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    # ファイル名の設定（現在の日時を含む）
    current_time = datetime.now().strftime("%Y年%m月%d日_%H時%M分%S秒")
    file_name = f"{target_username}_投稿データ_{current_time}.xlsx"
    
    # ファイルの保存
    wb.save(file_name)
    logging.info(f"データをExcelファイルに保存しました: {file_name}")

    return file_name

if __name__ == "__main__":
    try:
        login_username = validate_input(input("Threadsのログインユーザー名を入力してください: ").strip(), "ユーザー名")
        login_password = validate_input(input("Threadsのパスワードを入力してください: ").strip(), "パスワード")
        target_username = validate_input(input("スクレイピングしたいユーザー名を入力してください: ").strip(), "ターゲットユーザー名")

        driver = setup_driver()

        shared_data = {'stop': False, 'terminate': False}
        
        # キー入力リスナーをバックグラウンドで開始
        listener_thread = threading.Thread(target=key_listener, args=(shared_data,))
        listener_thread.start()
        
        if login_to_threads(driver, login_username, login_password):
            post_hrefs = access_threads_profile(driver, target_username, login_username, login_password)
            
            logging.info(f"収集された一意の投稿リンクの総数: {len(post_hrefs)} 件")
            logging.info("収集されたすべてのリンク:")
            for href in post_hrefs:
                print(href)

             # 投稿の処理
            processed_posts = process_posts(driver, post_hrefs, target_username)
            
            # キー入力リスナーを停止
            shared_data['stop'] = True
            listener_thread.join()
            
            logging.info(f"処理された投稿の総数: {len(processed_posts)} 件")

            # Excelファイルに保存
            excel_file = save_to_excel(processed_posts, target_username)
            print(f"データが {excel_file} に保存されました。")
            
        else:
            logging.error("ログインに失敗したため、スクレイピングを実行できません。")

    except ValueError as ve:
        logging.error(f"入力エラー: {ve}")
    except Exception as e:
        logging.error(f"予期せぬエラーが発生しました: {e}")
    finally:
        if 'driver' in locals() and driver:
            driver.quit()
        logging.info("すべてのリソースを解放しました。プログラムを終了します。")